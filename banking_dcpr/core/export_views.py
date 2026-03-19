# core/export_views.py
from django.http import HttpResponse
from django.db.models import Sum
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from datetime import timedelta
from decimal import Decimal

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from .models import PettyCashFund, PettyCashTransaction, BankAccount

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pcf_export_excel(request):
    """
    Export PCF data to Excel.
    Query params:
    - type: daily, weekly, monthly, unreplenished
    - date: YYYY-MM-DD (for daily)
    - start: YYYY-MM-DD (for weekly)
    - end: YYYY-MM-DD (for weekly)
    - year: YYYY (for monthly)
    - month: MM (for monthly)
    """
    export_type = request.query_params.get('type', 'daily')
    today = now().date()

    wb = Workbook()
    ws = wb.active
    ws.title = f"PCF {export_type.title()}"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = f"JOPCA PCF {export_type.title()} Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    row = 3
    ws[f'A{row}'] = "Generated:"
    ws[f'B{row}'] = today.strftime("%Y-%m-%d %H:%M:%S")
    row += 2

    headers = ['PCF Name', 'Location', 'Beginning', 'Disbursements', 'Replenishments', 'Unreplenished', 'Ending']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row += 1

    if export_type == 'daily':
        target_date_str = request.query_params.get('date')
        target_date = parse_date(target_date_str) if target_date_str else today
    elif export_type == 'weekly':
        start_str = request.query_params.get('start')
        end_str = request.query_params.get('end')
        start_date = parse_date(start_str) if start_str else today - timedelta(days=7)
        end_date = parse_date(end_str) if end_str else today
    elif export_type == 'monthly':
        year = request.query_params.get('year', today.year)
        month = request.query_params.get('month', today.month)
    else:
        target_date = today

    pcfs = PettyCashFund.objects.filter(is_active=True)
    grand_totals = {'beginning': 0, 'disbursements': 0, 'replenishments': 0, 'unreplenished': 0, 'ending': 0}

    for pcf in pcfs:
        if export_type == 'daily':
            beginning = pcf.opening_balance
            previous_txns = PettyCashTransaction.objects.filter(pcf=pcf, date__lt=target_date)
            for t in previous_txns:
                if t.type == 'disbursement':
                    beginning -= t.amount
                elif t.type == 'replenishment':
                    beginning += t.amount
            
            daily_txns = PettyCashTransaction.objects.filter(pcf=pcf, date=target_date)
            disbursements = daily_txns.filter(type='disbursement').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            replenishments = daily_txns.filter(type='replenishment').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            # Cumulative unreplenished = Total Disbursements - Total Replenishments
            total_disb = PettyCashTransaction.objects.filter(pcf=pcf, date__lte=target_date, type='disbursement').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            total_rep = PettyCashTransaction.objects.filter(pcf=pcf, date__lte=target_date, type='replenishment').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            unreplenished = max(Decimal('0'), total_disb - total_rep)
            ending = beginning - disbursements + replenishments
        else:
            beginning = pcf.opening_balance
            ending = pcf.current_balance
            disbursements = Decimal('0')
            replenishments = Decimal('0')
            total_disb = PettyCashTransaction.objects.filter(pcf=pcf, type='disbursement').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            total_rep = PettyCashTransaction.objects.filter(pcf=pcf, type='replenishment').aggregate(total=Sum('amount'))['total'] or Decimal('0')
            unreplenished = max(Decimal('0'), total_disb - total_rep)

        row_data = [
            pcf.name,
            pcf.get_location_display(),
            float(beginning),
            float(disbursements),
            float(replenishments),
            float(unreplenished),
            float(ending),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='left')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        grand_totals['beginning'] += beginning
        grand_totals['disbursements'] += disbursements
        grand_totals['replenishments'] += replenishments
        grand_totals['unreplenished'] += unreplenished
        grand_totals['ending'] += ending
        row += 1

    totals_row = [
        'GRAND TOTAL', '',
        float(grand_totals['beginning']),
        float(grand_totals['disbursements']),
        float(grand_totals['replenishments']),
        float(grand_totals['unreplenished']),
        float(grand_totals['ending']),
    ]

    totals_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, value in enumerate(totals_row, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = border
        cell.fill = totals_fill
        cell.font = Font(bold=True)
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="PCF_{export_type}_{today.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pcf_export_pdf(request):
    """
    Export PCF data to PDF.
    """
    export_type = request.query_params.get('type', 'daily')
    today = now().date()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="PCF_{export_type}_{today.strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"<b>JOPCA PCF {export_type.title()} Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(f"Generated: {today.strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    pcfs = PettyCashFund.objects.filter(is_active=True)

    data = [['PCF Name', 'Location', 'Beginning', 'Disbursements', 'Replenishments', 'Unreplenished', 'Ending']]

    grand_totals = {'beginning': 0, 'disbursements': 0, 'replenishments': 0, 'unreplenished': 0, 'ending': 0}

    for pcf in pcfs:
        beginning = float(pcf.opening_balance)
        
        total_disb = PettyCashTransaction.objects.filter(pcf=pcf, type='disbursement').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_rep = PettyCashTransaction.objects.filter(pcf=pcf, type='replenishment').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        disbursements = float(total_disb)
        replenishments = float(total_rep)
        unreplenished = max(0, disbursements - replenishments)
        ending = beginning - disbursements + replenishments

        data.append([
            pcf.name,
            pcf.get_location_display(),
            f"₱{beginning:,.2f}",
            f"₱{disbursements:,.2f}",
            f"₱{replenishments:,.2f}",
            f"₱{unreplenished:,.2f}",
            f"₱{ending:,.2f}",
        ])

        grand_totals['beginning'] += beginning
        grand_totals['disbursements'] += disbursements
        grand_totals['replenishments'] += replenishments
        grand_totals['unreplenished'] += unreplenished
        grand_totals['ending'] += ending

    data.append([
        'GRAND TOTAL', '',
        f"₱{grand_totals['beginning']:,.2f}",
        f"₱{grand_totals['disbursements']:,.2f}",
        f"₱{grand_totals['replenishments']:,.2f}",
        f"₱{grand_totals['unreplenished']:,.2f}",
        f"₱{grand_totals['ending']:,.2f}",
    ])

    table = Table(data, colWidths=[1.8*inch, 1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)

    doc.build(elements)
    return response
